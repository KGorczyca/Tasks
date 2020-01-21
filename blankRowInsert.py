'''Program wstawiający pusty wiersz. Program pobiera dwie liczby od użytkownika określane jako numN oraz numM. 
Począwszy od wiersza N, program powinien wstawić w arkuszu kalkulacyjnym M pustych wierszy'''

import openpyxl, os
wb = openpyxl.load_workbook('produceSales.xlsx')

wb.create_sheet(index=1, title = 'myProduce')
sheet1 = wb.get_sheet_by_name('Sheet')
sheet2 = wb.get_sheet_by_name('myProduce')

numN = int(input('Please enter your number N'))
numM = int(input('Please enter your number M'))

for i in range(1,int(numN)+1):
    for j in range(1,sheet1.max_column+1):
        sheet2.cell(row = i, column = j).value =  sheet1.cell(row = i, column = j).value
for i in range (numN+1, sheet1.max_row+1):
    for j in range(1, sheet1.max_column+1):
        sheet2.cell(row = i+numM, column = j).value = sheet1.cell(row = i, column = j).value
        

wb.save('produceSales.xlsx')

