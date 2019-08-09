'''program otwiera arkusz kalkulacyjny, a następnie
komórki kolumny A umieszcza w jednym pliku tekstowym, komórki
kolumny B w nastepnym itd.'''

import openpyxl, os

wb = openpyxl.load_workbook('FILE3.xlsx')

sheet = wb.get_active_sheet()




for i in range(1,sheet.max_column+1):
    text = open('new_file'+str(i)+'.txt', 'w') #otwiera nowy dokument dla kolejnej kolumny
    for j in range(1, sheet.max_row+1):
        text.write(str(sheet.cell(row = j, column = i).value)) #zapisuje wartość komórek kolumny w uprzednio utworzonym dokumencie

text.close()

