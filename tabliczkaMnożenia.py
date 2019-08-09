'''Program tworzący tabliczkę mnożenia w arkuszu kalkulacyjnym Excela.
Pierwszy wiersz i kolumna powinny być używane w charakterze etykiet
i dlatego znajdujący się w nich tekst ma być wyszczególniony.'''


import openpyxl
from openpyxl.styles import Font, NamedStyle, PatternFill,Side, Color,Border

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
sheet.title = 'multi_tabl'

highlight = NamedStyle(name = 'highlight')
highlight.font = Font(bold = True, size = 24, italic = True, color = 'FFBB00', name = 'Gigi')
bd = Side(style = 'double', color = 'FFBB00')
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.fill = PatternFill( 'solid', start_color = 'FFFF0000', end_color = 'FF0000FF')
highlight.color = Color(theme = 6, tint = 0.5)
wb.add_named_style(highlight)
        
for i in range(1,11):
    sheet['A' + str(i+1)] = i
    sheet.cell(row = 1, column = i+1).value = i
    sheet['A'+str(i+1)].style = highlight
    sheet.cell(row = 1, column = i+1).style = highlight

for i in range (1,11):
    for j in range(1,11):
        sheet.cell(row = i+1, column = j+1).value = i*j

wb.save('multiplication_table.xlsx')

